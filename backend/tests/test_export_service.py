import base64
import csv
import io
import unittest
import unittest.mock
import zipfile
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.schemas.export import DATEV_LIMITATIONS, EXPORT_COLUMNS, EXPORT_FORMAT_VERSION, ExportFormat
from app.schemas.invoice import (
    InvoiceParseResponse,
    InvoiceTotals,
    LineItem,
    MismatchField,
    PartyInfo,
    ParseStatus,
    ValidationIssue,
    ValidationStatus,
)
from app.helper_functions.filenames import safe_filename_stem
from app.services.export_service import (
    ExportService,
    build_datev_row,
    build_export_filename,
    build_flat_rows,
    build_package_filename,
    build_package_summary,
)
from app.services.validation_report import (
    build_validation_report,
    build_validation_report_filename,
)


def _sample_invoice() -> InvoiceParseResponse:
    return InvoiceParseResponse(
        status=ParseStatus.SUCCESS,
        message="ok",
        filename="sample.xml",
        file_type="xrechnung_xml",
        invoice_number="2025/10294",
        issue_date="2025-01-31",
        due_date="2025-01-31",
        seller=PartyInfo(
            name="KMLZ Rechtsanwaltsges. mbH",
            vat_id="DE814742004",
            iban="DE95700400410228840500",
        ),
        buyer=PartyInfo(name="Buyer AG", vat_id="DE123"),
        totals=InvoiceTotals(net=227.5, tax=43.23, gross=270.73, currency="EUR"),
        line_items=[
            LineItem(
                position=1,
                description="Beratung",
                quantity=1,
                unit_price=227.5,
                tax_rate=19.0,
                net_amount=227.5,
            )
        ],
        payment_reference="REF-1",
    )


class TestExportService(unittest.TestCase):
    def setUp(self) -> None:
        self.service: ExportService = ExportService()
        self.invoice: InvoiceParseResponse = _sample_invoice()

    def test_filename_convention(self) -> None:
        name: str = build_export_filename(self.invoice, ExportFormat.CSV)
        self.assertTrue(name.endswith(".csv"))
        self.assertIn("2025", name)
        self.assertIn("20250131", name)
        self.assertIn("KMLZ", name)

    def test_csv_columns_and_row(self) -> None:
        content, media, filename = self.service.export(self.invoice, ExportFormat.CSV)
        self.assertIn("text/csv", media)
        self.assertTrue(filename.endswith(".csv"))
        text: str = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text), delimiter=";")
        self.assertEqual(list(reader.fieldnames or []), EXPORT_COLUMNS)
        rows = list(reader)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["invoice_number"], "2025/10294")
        self.assertEqual(rows[0]["gross"], "270,73")
        self.assertEqual(rows[0]["issue_date"], "31.01.2025")
        self.assertEqual(rows[0]["line_description"], "Beratung")

    def test_excel_sheets(self) -> None:
        content, media, filename = self.service.export(self.invoice, ExportFormat.EXCEL)
        self.assertIn("spreadsheetml", media)
        self.assertTrue(filename.endswith(".xlsx"))
        workbook = load_workbook(io.BytesIO(content))
        self.assertEqual(set(workbook.sheetnames), {"Invoice", "Lines", "Flat"})
        self.assertGreaterEqual(workbook["Lines"].max_row, 2)
        self.assertEqual(workbook["Invoice"]["A2"].value, "export_format_version")
        self.assertEqual(workbook["Invoice"]["B2"].value, EXPORT_FORMAT_VERSION)

    def test_datev_german_decimal_and_encoding(self) -> None:
        content, media, filename = self.service.export(self.invoice, ExportFormat.DATEV)
        self.assertTrue(filename.startswith("datev_"))
        text: str = content.decode("cp1252")
        self.assertIn("270,73", text)
        self.assertIn("2025/10294", text)
        self.assertIn("S;", text.replace("\r\n", "\n") or "S")
        self.assertIn("31012025", text)

    def test_datev_credit_note_uses_haben(self) -> None:
        invoice: InvoiceParseResponse = _sample_invoice()
        invoice.document_type = "credit_note"
        row = build_datev_row(invoice)
        self.assertEqual(row["Soll/Haben-Kennzeichen"], "H")
        self.assertEqual(row["Umsatz"], "270,73")

    def test_safe_filename_transliterates_umlauts(self) -> None:
        self.assertEqual(safe_filename_stem("Müller GmbH"), "Mueller_GmbH")
        invoice: InvoiceParseResponse = _sample_invoice()
        invoice.seller = PartyInfo(name="Müller & Söhne")
        name: str = build_export_filename(invoice, ExportFormat.CSV)
        self.assertIn("Mueller_Soehne", name)
        self.assertNotIn("ü", name)

    def test_empty_optional_fields_safe(self) -> None:
        invoice: InvoiceParseResponse = InvoiceParseResponse(
            status=ParseStatus.PARTIAL,
            message="partial",
            filename="x.xml",
            invoice_number="1",
            totals=InvoiceTotals(gross=10.0, currency="EUR"),
            line_items=[],
        )
        rows = build_flat_rows(invoice)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["seller_name"], "")
        self.assertEqual(rows[0]["line_description"], "")

    def test_accountant_package_zip_contents(self) -> None:
        pdf_bytes: bytes = b"%PDF-1.4 minimal test pdf"
        xml_bytes: bytes = b"<?xml version='1.0'?><Invoice/>"
        content, media, filename = self.service.build_accountant_package(
            invoice=self.invoice,
            pdf_bytes=pdf_bytes,
            pdf_filename="Beleg_Test.PDF",
            xml_bytes=xml_bytes,
            xml_filename="rechnung.xml",
        )
        self.assertEqual(media, "application/zip")
        self.assertTrue(filename.startswith("buchhaltung_"))
        self.assertTrue(filename.endswith(".zip"))

        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names: list[str] = sorted(archive.namelist())
            self.assertIn("summary.txt", names)
            self.assertIn("export_manifest.txt", names)
            self.assertIn("datev_hinweise.txt", names)
            self.assertTrue(any(name.startswith("pruefbericht_") for name in names))
            self.assertTrue(any(name.endswith(".xlsx") for name in names))
            self.assertTrue(any(name.startswith("datev_") for name in names))
            self.assertIn("original/Beleg_Test.pdf", names)
            self.assertIn("original/rechnung.xml", names)
            summary: str = archive.read("summary.txt").decode("utf-8")
            self.assertIn("2025/10294", summary)
            self.assertIn("270,73", summary)
            self.assertIn(EXPORT_FORMAT_VERSION, summary)
            manifest: str = archive.read("export_manifest.txt").decode("utf-8")
            self.assertIn(EXPORT_FORMAT_VERSION, manifest)
            notes: str = archive.read("datev_hinweise.txt").decode("utf-8")
            self.assertIn("kein DATEVconnect", notes)
            self.assertIn("DATEVconnect", DATEV_LIMITATIONS)

    def test_accountant_package_extracts_xml_from_zugferd_pdf(self) -> None:
        pdf_bytes: bytes = b"%PDF-1.4 placeholder"
        xml_text: str = "<?xml version='1.0'?><rsm:CrossIndustryInvoice/>"
        with unittest.mock.patch(
            "app.services.export_service.extract_embedded_xml_from_pdf",
            return_value=xml_text,
        ):
            content, _, _ = self.service.build_accountant_package(
                invoice=self.invoice,
                pdf_bytes=pdf_bytes,
                pdf_filename="zugferd.pdf",
            )
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names: list[str] = archive.namelist()
            self.assertIn("original/zugferd.pdf", names)
            xml_members: list[str] = [name for name in names if name.endswith(".xml")]
            self.assertEqual(len(xml_members), 1)
            self.assertEqual(archive.read(xml_members[0]).decode("utf-8"), xml_text)

    def test_package_summary_mismatch_line(self) -> None:
        invoice: InvoiceParseResponse = _sample_invoice()
        invoice.file_type = "zugferd_pdf"
        invoice.validation_status = ValidationStatus.WARNING
        invoice.mismatch_fields = [
            MismatchField(
                field="gross",
                label="Brutto",
                xml_value="100.00",
                pdf_value="99.00",
                matched=False,
            )
        ]
        text: str = build_package_summary(invoice)
        self.assertIn("Abweichung (Brutto)", text)
        self.assertEqual(
            build_package_filename(invoice),
            "buchhaltung_KMLZ_Rechtsanwaltsges_mbH_2025_10294_20250131.zip",
        )

    def test_validation_report_for_mismatch(self) -> None:
        invoice: InvoiceParseResponse = _sample_invoice()
        invoice.file_type = "zugferd_pdf"
        invoice.validation_status = ValidationStatus.INVALID
        invoice.validation_issues = [
            ValidationIssue(
                level="error",
                category="mismatch",
                code="MISMATCH_IBAN",
                message="IBAN in PDF und XML weicht ab.",
                field="iban",
            )
        ]
        invoice.mismatch_fields = [
            MismatchField(
                field="iban",
                label="IBAN",
                xml_value="DE95700400410228840500",
                pdf_value="DE00",
                matched=False,
            )
        ]
        invoice.next_steps = [
            "Nicht zahlen. PDF und XML weichen ab — Lieferanten kontaktieren."
        ]
        text: str = build_validation_report(invoice)
        self.assertIn("Prüfbericht", text)
        self.assertIn("Nicht zahlen", text)
        self.assertIn("KMLZ", text)
        self.assertIn("Abweichung", text)
        self.assertTrue(
            build_validation_report_filename(invoice).startswith("pruefbericht_")
        )
        self.assertTrue(build_validation_report_filename(invoice).endswith(".txt"))


class TestExportApi(unittest.TestCase):
    def setUp(self) -> None:
        self.client: TestClient = TestClient(app)

    def test_export_csv_endpoint(self) -> None:
        response = self.client.post(
            "/api/invoices/export",
            json={"format": "csv", "invoice": _sample_invoice().model_dump(mode="json")},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response.headers["content-type"])
        self.assertIn("attachment", response.headers.get("content-disposition", ""))

    def test_export_mapping_docs(self) -> None:
        response = self.client.get("/api/invoices/export/mapping")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertGreaterEqual(len(data), 3)
        versions: set[str] = {item["version"] for item in data}
        self.assertEqual(versions, {EXPORT_FORMAT_VERSION})
        datev = next(item for item in data if item["format"] == "datev")
        self.assertIn("DATEVconnect", datev["limitations"])
        self.assertIn("kein DATEVconnect", datev["limitations"])
        self.assertEqual(datev["encoding"], "cp1252")
        csv_doc = next(item for item in data if item["format"] == "csv")
        self.assertEqual(csv_doc["delimiter"], ";")
        self.assertEqual(csv_doc["decimal_separator"], ",")

    def test_parse_then_export_roundtrip(self) -> None:
        fixtures: Path = Path(__file__).parent / "xml_files" / "xml_text_from_zugpferd.xml"
        parse_response = self.client.post(
            "/api/invoices/parse",
            files={"file": ("sample.xml", fixtures.read_bytes(), "application/xml")},
        )
        self.assertEqual(parse_response.status_code, 200)
        invoice = parse_response.json()
        export_response = self.client.post(
            "/api/invoices/export",
            json={"format": "excel", "invoice": invoice},
        )
        self.assertEqual(export_response.status_code, 200)
        self.assertTrue(len(export_response.content) > 100)

    def test_accountant_package_endpoint(self) -> None:
        pdf_b64: str = base64.b64encode(b"%PDF-1.4 package").decode("ascii")
        xml_b64: str = base64.b64encode(b"<?xml version='1.0'?><Invoice/>").decode("ascii")
        response = self.client.post(
            "/api/invoices/export/accountant-package",
            json={
                "invoice": _sample_invoice().model_dump(mode="json"),
                "pdf_base64": pdf_b64,
                "pdf_filename": "original.pdf",
                "xml_base64": xml_b64,
                "xml_filename": "original.xml",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/zip", response.headers["content-type"])
        self.assertIn("buchhaltung_", response.headers.get("content-disposition", ""))
        with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
            names: list[str] = archive.namelist()
            self.assertIn("summary.txt", names)
            self.assertIn("original/original.pdf", names)
            self.assertIn("original/original.xml", names)
            self.assertIn("export_manifest.txt", names)
            self.assertIn("datev_hinweise.txt", names)

    def test_accountant_package_rejects_invalid_xml(self) -> None:
        response = self.client.post(
            "/api/invoices/export/accountant-package",
            json={
                "invoice": _sample_invoice().model_dump(mode="json"),
                "xml_base64": base64.b64encode(b"not-xml").decode("ascii"),
            },
        )
        self.assertEqual(response.status_code, 400)

    def test_validation_report_endpoint_allows_invalid_invoice(self) -> None:
        invoice: InvoiceParseResponse = _sample_invoice()
        invoice.status = ParseStatus.PARTIAL
        invoice.validation_status = ValidationStatus.INVALID
        invoice.validation_issues = [
            ValidationIssue(
                level="error",
                category="business",
                code="BT-1_MISSING",
                message="Rechnungsnummer fehlt.",
            )
        ]
        response = self.client.post(
            "/api/invoices/export/validation-report",
            json={"invoice": invoice.model_dump(mode="json")},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/plain", response.headers["content-type"])
        self.assertIn("pruefbericht_", response.headers.get("content-disposition", ""))
        body: str = response.content.decode("utf-8-sig")
        self.assertIn("Prüfbericht", body)
        self.assertIn("ungültig", body)

    def test_error_invoice_cannot_be_exported_as_accounting_file(self) -> None:
        invoice: InvoiceParseResponse = InvoiceParseResponse(
            status=ParseStatus.ERROR,
            message="unreadable",
            filename="bad.xml",
            validation_status=ValidationStatus.INVALID,
        )
        response = self.client.post(
            "/api/invoices/export",
            json={"format": "csv", "invoice": invoice.model_dump(mode="json")},
        )
        self.assertEqual(response.status_code, 400)


if __name__ == "__main__":
    unittest.main()
