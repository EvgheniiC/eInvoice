import base64
import binascii
import csv
import io
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.helper_functions.filenames import safe_filename_stem
from app.schemas.export import (
    DATEV_COLUMNS,
    DATEV_LIMITATIONS,
    EXPORT_COLUMNS,
    EXPORT_FORMAT_VERSION,
    ExportFormat,
)
from app.schemas.invoice import (
    InvoiceParseResponse,
    LineItem,
    MismatchField,
    ParseStatus,
    ValidationMeta,
)
from app.services.pdf_xml_extractor import extract_embedded_xml_from_pdf
from app.services.validation_report import (
    build_validation_report,
    build_validation_report_filename,
)

MAX_PACKAGE_SOURCE_BYTES: int = 50 * 1024 * 1024
ORIGINAL_DIR: str = "original"


@dataclass(frozen=True)
class BatchPackageEntry:
    """One batch file: original bytes (if still on disk) plus optional parse DTO."""

    filename: str
    original_bytes: Optional[bytes]
    invoice: Optional[InvoiceParseResponse]


class ExportService:
    """Build CSV / Excel / DATEV files from the public invoice DTO."""

    def export(self, invoice: InvoiceParseResponse, export_format: ExportFormat) -> Tuple[bytes, str, str]:
        """
        Return (file_bytes, media_type, download_filename).
        """
        filename: str = build_export_filename(invoice=invoice, export_format=export_format)

        if export_format == ExportFormat.CSV:
            return self._to_csv(invoice), "text/csv; charset=utf-8", filename
        if export_format == ExportFormat.EXCEL:
            return (
                self._to_excel(invoice),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename,
            )
        if export_format == ExportFormat.DATEV:
            return self._to_datev(invoice), "text/csv; charset=cp1252", filename
        raise ValueError(f"Unsupported export format: {export_format}")

    def build_accountant_package(
        self,
        invoice: InvoiceParseResponse,
        pdf_bytes: Optional[bytes] = None,
        pdf_filename: Optional[str] = None,
        xml_bytes: Optional[bytes] = None,
        xml_filename: Optional[str] = None,
    ) -> Tuple[bytes, str, str]:
        """
        ZIP for Steuerberater: original + summary + Prüfbericht + Excel + DATEV.
        Returns (zip_bytes, media_type, download_filename).
        """
        _assert_source_size(pdf_bytes, "PDF")
        _assert_source_size(xml_bytes, "XML")

        extracted_xml: Optional[bytes] = None
        if xml_bytes is None and pdf_bytes is not None:
            extracted_xml = _extract_xml_from_zugferd(pdf_bytes)

        original_xml: Optional[bytes] = xml_bytes if xml_bytes is not None else extracted_xml
        has_pdf: bool = pdf_bytes is not None
        has_xml: bool = original_xml is not None

        buffer: io.BytesIO = io.BytesIO()
        with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(
                "export_manifest.txt",
                build_export_manifest(invoice, has_xml=has_xml, has_pdf=has_pdf),
            )
            archive.writestr("datev_hinweise.txt", build_datev_notes())
            archive.writestr(
                "summary.txt",
                build_package_summary(invoice, has_xml=has_xml, has_pdf=has_pdf),
            )
            archive.writestr(
                build_validation_report_filename(invoice),
                build_validation_report(invoice),
            )

            excel_bytes, _, excel_name = self.export(invoice, ExportFormat.EXCEL)
            archive.writestr(excel_name, excel_bytes)

            datev_bytes, _, datev_name = self.export(invoice, ExportFormat.DATEV)
            archive.writestr(datev_name, datev_bytes)

            if original_xml is not None:
                archive.writestr(
                    _package_source_member_name(invoice, xml_filename, "xml"),
                    original_xml,
                )
            if pdf_bytes is not None:
                archive.writestr(
                    _package_source_member_name(invoice, pdf_filename, "pdf"),
                    pdf_bytes,
                )

        zip_name: str = build_package_filename(invoice)
        return buffer.getvalue(), "application/zip", zip_name

    def build_batch_accountant_package(
        self,
        entries: List[BatchPackageEntry],
        completed_at: datetime,
    ) -> Tuple[bytes, str, str]:
        """
        One ZIP for N invoices: combined Excel + DATEV + manifest + originals.
        Returns (zip_bytes, media_type, download_filename).
        """
        exportable: List[InvoiceParseResponse] = [
            entry.invoice
            for entry in entries
            if entry.invoice is not None and invoice_is_exportable(entry.invoice)
        ]
        if not exportable:
            raise ValueError("Keine exportierbare Rechnung in diesem Auftrag.")

        originals: List[tuple[str, bytes]] = []
        for index, entry in enumerate(entries, start=1):
            if entry.original_bytes is None:
                continue
            _assert_source_size(entry.original_bytes, "Originaldatei")
            originals.append(
                (
                    _batch_original_member_name(index, entry.filename),
                    entry.original_bytes,
                )
            )

        date_part: str = completed_at.strftime("%Y%m%d")
        excel_name: str = f"rechnungen_{date_part}_{len(exportable)}.xlsx"
        datev_name: str = f"datev_rechnungen_{date_part}_{len(exportable)}.csv"
        has_xml: bool = any(name.lower().endswith(".xml") for name, _ in originals)
        has_pdf: bool = any(name.lower().endswith(".pdf") for name, _ in originals)

        buffer: io.BytesIO = io.BytesIO()
        with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(
                "export_manifest.txt",
                build_batch_export_manifest(
                    exportable,
                    original_names=[name for name, _ in originals],
                    has_xml=has_xml,
                    has_pdf=has_pdf,
                ),
            )
            archive.writestr("datev_hinweise.txt", build_datev_notes())
            archive.writestr(
                "summary.txt",
                build_batch_package_summary(
                    exportable,
                    entries=entries,
                    has_xml=has_xml,
                    has_pdf=has_pdf,
                ),
            )
            archive.writestr(
                "pruefbericht_paket.txt",
                build_batch_validation_report(exportable),
            )
            archive.writestr(excel_name, self._to_excel_many(exportable))
            archive.writestr(datev_name, self._to_datev_many(exportable))
            for member_name, payload in originals:
                archive.writestr(member_name, payload)

        zip_name: str = build_batch_package_filename(completed_at, len(exportable))
        return buffer.getvalue(), "application/zip", zip_name

    def _to_csv(self, invoice: InvoiceParseResponse) -> bytes:
        buffer: io.StringIO = io.StringIO()
        writer: csv.DictWriter = csv.DictWriter(
            buffer,
            fieldnames=EXPORT_COLUMNS,
            delimiter=";",
            extrasaction="ignore",
            lineterminator="\r\n",
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writeheader()
        for row in build_flat_rows(invoice):
            writer.writerow(_localize_csv_row(row))
        # UTF-8 BOM helps Excel on Windows open umlauts correctly
        return ("\ufeff" + buffer.getvalue()).encode("utf-8")

    def _to_excel(self, invoice: InvoiceParseResponse) -> bytes:
        workbook: Workbook = Workbook()
        header_sheet: Worksheet = workbook.active
        header_sheet.title = "Invoice"

        header_rows: List[Tuple[str, object]] = [
            ("export_format_version", EXPORT_FORMAT_VERSION),
            ("invoice_number", invoice.invoice_number or ""),
            ("issue_date", _de_date(invoice.issue_date)),
            ("due_date", _de_date(invoice.due_date)),
            ("seller_name", invoice.seller.name if invoice.seller else ""),
            ("seller_vat_id", invoice.seller.vat_id if invoice.seller else ""),
            ("seller_iban", invoice.seller.iban if invoice.seller else ""),
            ("buyer_name", invoice.buyer.name if invoice.buyer else ""),
            ("buyer_vat_id", invoice.buyer.vat_id if invoice.buyer else ""),
            ("currency", invoice.totals.currency if invoice.totals else ""),
            ("net", invoice.totals.net if invoice.totals else ""),
            ("tax", invoice.totals.tax if invoice.totals else ""),
            ("gross", invoice.totals.gross if invoice.totals else ""),
            ("payment_reference", invoice.payment_reference or ""),
        ]
        header_sheet.append(["field", "value"])
        for key, value in header_rows:
            header_sheet.append([key, value if value is not None else ""])

        lines_sheet: Worksheet = workbook.create_sheet("Lines")
        lines_sheet.append(
            [
                "position",
                "description",
                "quantity",
                "unit",
                "unit_price",
                "tax_rate",
                "net_amount",
                "gross_amount",
            ]
        )
        for item in invoice.line_items:
            lines_sheet.append(
                [
                    item.position,
                    item.description or "",
                    item.quantity,
                    item.unit or "",
                    item.unit_price,
                    item.tax_rate,
                    item.net_amount,
                    item.gross_amount,
                ]
            )

        flat_sheet: Worksheet = workbook.create_sheet("Flat")
        flat_sheet.append(EXPORT_COLUMNS)
        for row in build_flat_rows(invoice):
            localized: Dict[str, Any] = _localize_csv_row(row)
            flat_sheet.append([localized.get(col, "") for col in EXPORT_COLUMNS])

        output: io.BytesIO = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _to_datev(self, invoice: InvoiceParseResponse) -> bytes:
        """
        Minimal DATEV Buchungsstapel CSV:
        - semicolon separator
        - German decimal comma
        - CP1252 encoding
        - one Soll/Haben line for gross amount
        """
        buffer: io.StringIO = io.StringIO()
        writer: csv.DictWriter = csv.DictWriter(
            buffer,
            fieldnames=DATEV_COLUMNS,
            delimiter=";",
            extrasaction="ignore",
            lineterminator="\r\n",
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writeheader()
        writer.writerow(build_datev_row(invoice))
        return buffer.getvalue().encode("cp1252", errors="replace")

    def _to_excel_many(self, invoices: List[InvoiceParseResponse]) -> bytes:
        workbook: Workbook = Workbook()
        header_sheet: Worksheet = workbook.active
        header_sheet.title = "Invoice"
        header_sheet.append(["export_format_version", EXPORT_FORMAT_VERSION])
        header_sheet.append(["invoice_count", len(invoices)])
        header_sheet.append([])
        header_sheet.append(
            [
                "invoice_number",
                "issue_date",
                "due_date",
                "seller_name",
                "seller_vat_id",
                "seller_iban",
                "buyer_name",
                "currency",
                "net",
                "tax",
                "gross",
                "payment_reference",
                "filename",
                "validation_status",
            ]
        )
        for invoice in invoices:
            header_sheet.append(
                [
                    invoice.invoice_number or "",
                    _de_date(invoice.issue_date) or invoice.issue_date or "",
                    _de_date(invoice.due_date) or invoice.due_date or "",
                    invoice.seller.name if invoice.seller and invoice.seller.name else "",
                    invoice.seller.vat_id if invoice.seller and invoice.seller.vat_id else "",
                    invoice.seller.iban if invoice.seller and invoice.seller.iban else "",
                    invoice.buyer.name if invoice.buyer and invoice.buyer.name else "",
                    invoice.totals.currency if invoice.totals and invoice.totals.currency else "",
                    invoice.totals.net if invoice.totals else "",
                    invoice.totals.tax if invoice.totals else "",
                    invoice.totals.gross if invoice.totals else "",
                    invoice.payment_reference or "",
                    invoice.filename,
                    invoice.validation_status.value,
                ]
            )

        lines_sheet: Worksheet = workbook.create_sheet("Lines")
        lines_sheet.append(
            [
                "invoice_number",
                "position",
                "description",
                "quantity",
                "unit",
                "unit_price",
                "tax_rate",
                "net_amount",
                "gross_amount",
            ]
        )
        for invoice in invoices:
            number: str = invoice.invoice_number or ""
            for item in invoice.line_items:
                lines_sheet.append(
                    [
                        number,
                        item.position,
                        item.description or "",
                        item.quantity,
                        item.unit or "",
                        item.unit_price,
                        item.tax_rate,
                        item.net_amount,
                        item.gross_amount,
                    ]
                )

        flat_sheet: Worksheet = workbook.create_sheet("Flat")
        flat_sheet.append(EXPORT_COLUMNS)
        for invoice in invoices:
            for row in build_flat_rows(invoice):
                localized: Dict[str, Any] = _localize_csv_row(row)
                flat_sheet.append([localized.get(col, "") for col in EXPORT_COLUMNS])

        output: io.BytesIO = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _to_datev_many(self, invoices: List[InvoiceParseResponse]) -> bytes:
        buffer: io.StringIO = io.StringIO()
        writer: csv.DictWriter = csv.DictWriter(
            buffer,
            fieldnames=DATEV_COLUMNS,
            delimiter=";",
            extrasaction="ignore",
            lineterminator="\r\n",
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writeheader()
        for invoice in invoices:
            writer.writerow(build_datev_row(invoice))
        return buffer.getvalue().encode("cp1252", errors="replace")


def build_flat_rows(invoice: InvoiceParseResponse) -> List[Dict[str, Any]]:
    """One export row per line item; header fields repeated. Empty line if no positions."""
    base: Dict[str, Any] = {
        "invoice_number": invoice.invoice_number or "",
        "issue_date": invoice.issue_date or "",
        "due_date": invoice.due_date or "",
        "seller_name": invoice.seller.name if invoice.seller and invoice.seller.name else "",
        "seller_vat_id": invoice.seller.vat_id if invoice.seller and invoice.seller.vat_id else "",
        "seller_iban": invoice.seller.iban if invoice.seller and invoice.seller.iban else "",
        "buyer_name": invoice.buyer.name if invoice.buyer and invoice.buyer.name else "",
        "buyer_vat_id": invoice.buyer.vat_id if invoice.buyer and invoice.buyer.vat_id else "",
        "currency": invoice.totals.currency if invoice.totals and invoice.totals.currency else "",
        "net": _num_str(invoice.totals.net if invoice.totals else None),
        "tax": _num_str(invoice.totals.tax if invoice.totals else None),
        "gross": _num_str(invoice.totals.gross if invoice.totals else None),
        "payment_reference": invoice.payment_reference or "",
    }

    items: List[LineItem] = invoice.line_items
    if not items:
        row: Dict[str, Any] = dict(base)
        row.update(
            {
                "line_position": "",
                "line_description": "",
                "line_quantity": "",
                "line_unit": "",
                "line_unit_price": "",
                "line_tax_rate": "",
                "line_net_amount": "",
            }
        )
        return [row]

    rows: List[Dict[str, Any]] = []
    for item in items:
        row = dict(base)
        row.update(
            {
                "line_position": item.position if item.position is not None else "",
                "line_description": item.description or "",
                "line_quantity": _num_str(item.quantity),
                "line_unit": item.unit or "",
                "line_unit_price": _num_str(item.unit_price),
                "line_tax_rate": _num_str(item.tax_rate),
                "line_net_amount": _num_str(item.net_amount),
            }
        )
        rows.append(row)
    return rows


def build_datev_row(invoice: InvoiceParseResponse) -> Dict[str, str]:
    amount: Optional[Decimal] = invoice.totals.gross if invoice.totals else None
    seller_name: str = invoice.seller.name if invoice.seller and invoice.seller.name else "Lieferant"
    invoice_no: str = invoice.invoice_number or ""
    booking_text: str = f"{seller_name} {invoice_no}".strip()[:60]
    belegdatum: str = _datev_date(invoice.issue_date)
    currency: str = (
        invoice.totals.currency if invoice.totals and invoice.totals.currency else "EUR"
    )
    signed_amount: Decimal = amount if amount is not None else Decimal("0")
    is_credit: bool = invoice.document_type == "credit_note" or signed_amount < 0
    kennzeichen: str = "H" if is_credit else "S"

    return {
        "Umsatz": _de_amount(abs(signed_amount) if amount is not None else None),
        "Soll/Haben-Kennzeichen": kennzeichen,
        "WKZ Umsatz": currency,
        "Kurs": "",
        "Basis-Umsatz": "",
        "WKZ Basis-Umsatz": "",
        "Konto": "",
        "Gegenkonto (ohne BU-Schlüssel)": "",
        "BU-Schlüssel": "",
        "Belegdatum": belegdatum,
        "Belegfeld 1": invoice_no[:36],
        "Belegfeld 2": "",
        "Skonto": "",
        "Buchungstext": booking_text,
    }


def build_export_filename(invoice: InvoiceParseResponse, export_format: ExportFormat) -> str:
    """Filename convention: supplier_invoiceNo_date.ext"""
    supplier: str = safe_filename_stem(invoice.seller.name if invoice.seller else None) or "supplier"
    number: str = safe_filename_stem(invoice.invoice_number) or "invoice"
    date_part: str = (invoice.issue_date or "nodate").replace("-", "")
    extension: str = {
        ExportFormat.CSV: "csv",
        ExportFormat.EXCEL: "xlsx",
        ExportFormat.DATEV: "csv",
    }[export_format]
    prefix: str = "datev_" if export_format == ExportFormat.DATEV else ""
    return f"{prefix}{supplier}_{number}_{date_part}.{extension}"


def build_package_filename(invoice: InvoiceParseResponse) -> str:
    """ZIP filename: buchhaltung_supplier_invoiceNo_date.zip"""
    supplier: str = safe_filename_stem(invoice.seller.name if invoice.seller else None) or "supplier"
    number: str = safe_filename_stem(invoice.invoice_number) or "invoice"
    date_part: str = (invoice.issue_date or "nodate").replace("-", "")
    return f"buchhaltung_{supplier}_{number}_{date_part}.zip"


def build_batch_package_filename(completed_at: datetime, invoice_count: int) -> str:
    """ZIP filename: buchhaltung_paket_YYYYMMDD_NDateien.zip"""
    date_part: str = completed_at.strftime("%Y%m%d")
    return f"buchhaltung_paket_{date_part}_{invoice_count}Dateien.zip"


def invoice_is_exportable(invoice: InvoiceParseResponse) -> bool:
    """Same gate as the single-invoice export endpoints."""
    if invoice.status == ParseStatus.ERROR:
        return False
    if not invoice.invoice_number and not (invoice.totals and invoice.totals.gross):
        return False
    return True


def build_package_summary(
    invoice: InvoiceParseResponse,
    has_xml: bool = False,
    has_pdf: bool = False,
) -> str:
    """Short German summary for the accountant package."""
    currency: str = (
        invoice.totals.currency if invoice.totals and invoice.totals.currency else "EUR"
    )
    net: str = _de_amount(invoice.totals.net if invoice.totals else None) or "—"
    tax: str = _de_amount(invoice.totals.tax if invoice.totals else None) or "—"
    gross: str = _de_amount(invoice.totals.gross if invoice.totals else None) or "—"
    seller_name: str = invoice.seller.name if invoice.seller and invoice.seller.name else "—"
    seller_iban: str = invoice.seller.iban if invoice.seller and invoice.seller.iban else "—"
    buyer_name: str = invoice.buyer.name if invoice.buyer and invoice.buyer.name else "—"
    mismatch_line: str = _mismatch_summary_line(invoice)

    lines: List[str] = [
        "Buchhaltungspaket — eInvoice",
        "============================",
        "",
        f"Exportformat: {EXPORT_FORMAT_VERSION}",
        f"Rechnung: {invoice.invoice_number or '—'}",
        f"Datum: {_de_date(invoice.issue_date) or '—'}",
        f"Fälligkeitsdatum: {_de_date(invoice.due_date) or '—'}",
        f"Zahlungsreferenz: {invoice.payment_reference or '—'}",
        f"Quelldatei: {invoice.filename}",
        f"Typ: {invoice.file_type or '—'}",
        "",
        f"Lieferant: {seller_name}",
        f"IBAN: {seller_iban}",
        f"Empfänger: {buyer_name}",
        "",
        f"Netto: {net} {currency}",
        f"MwSt: {tax} {currency}",
        f"Brutto: {gross} {currency}",
        "",
        f"Parse-Status: {invoice.status.value}",
        f"Prüfung: {invoice.validation_status.value}",
        f"Standard: {invoice.validation_meta.standard_version or '—'}",
        f"Profil: {invoice.validation_meta.profile or '—'}",
        f"Prüfengine: {_validation_engine_line(invoice)}",
        f"PDF↔XML: {mismatch_line}",
        "",
        "Inhalt dieses ZIP:",
        "- export_manifest.txt (Formatversion und Dateiliste)",
        "- datev_hinweise.txt (DATEV-Grenzen, kein DATEVconnect)",
        "- summary.txt (diese Datei)",
        "- Prüfbericht (für Lieferant oder Steuerberater)",
        "- Excel-Export (Übersicht + Positionen)",
        "- DATEV-Export (Buchungsvorschlag-CSV)",
    ]
    if has_xml:
        lines.append("- original/*.xml (ursprüngliches Rechnungs-XML)")
    if has_pdf:
        lines.append("- original/*.pdf (ZUGFeRD-PDF mit eingebettetem XML)")
    if not has_xml and not has_pdf:
        lines.append("- Originaldatei fehlt in diesem Paket")

    lines.extend(
        [
            "",
            "Hinweis: Die Prüfung betrifft Schema-/Standardkonformität.",
            "Die Entscheidung über den Vorsteuerabzug liegt beim Steuerberater.",
            DATEV_LIMITATIONS,
        ]
    )

    if invoice.mismatch_warnings:
        lines.extend(["", "Abweichungen / Hinweise:"])
        for warning in invoice.mismatch_warnings:
            lines.append(f"- {warning}")

    if invoice.validation_issues:
        lines.extend(["", "Prüfungshinweise:"])
        for issue in invoice.validation_issues[:20]:
            bt: str = f" {issue.bt_code}" if issue.bt_code else ""
            lines.append(f"- [{issue.level}/{issue.category}{bt}] {issue.message}")
            if issue.explanation:
                lines.append(f"  {issue.explanation}")

    if invoice.next_steps:
        lines.extend(["", "Empfohlene nächste Schritte:"])
        for index, step in enumerate(invoice.next_steps, start=1):
            lines.append(f"{index}. {step}")

    lines.append("")
    return "\n".join(lines)


def build_batch_package_summary(
    invoices: List[InvoiceParseResponse],
    entries: List[BatchPackageEntry],
    has_xml: bool,
    has_pdf: bool,
) -> str:
    """German overview of all invoices in the batch accountant package."""
    original_count: int = sum(1 for entry in entries if entry.original_bytes is not None)
    skipped_count: int = len(entries) - len(invoices)
    lines: List[str] = [
        "Buchhaltungspaket — eInvoice (Sammelexport)",
        "==========================================",
        "",
        f"Exportformat: {EXPORT_FORMAT_VERSION}",
        f"Rechnungen im Export: {len(invoices)}",
        f"Dateien im Auftrag: {len(entries)}",
        f"Originaldateien im ZIP: {original_count}",
        "",
        "Rechnungen:",
    ]
    for invoice in invoices:
        currency: str = (
            invoice.totals.currency if invoice.totals and invoice.totals.currency else "EUR"
        )
        gross: str = _de_amount(invoice.totals.gross if invoice.totals else None) or "—"
        seller_name: str = invoice.seller.name if invoice.seller and invoice.seller.name else "—"
        lines.append(
            f"- {invoice.invoice_number or '—'} · {_de_date(invoice.issue_date) or '—'} · "
            f"{seller_name} · {gross} {currency} · Prüfung: {invoice.validation_status.value} · "
            f"{invoice.filename}"
        )
    if skipped_count > 0:
        lines.extend(
            [
                "",
                f"{skipped_count} Datei(en) ohne exportierbare Rechnung sind nicht in Excel/DATEV.",
            ]
        )
    lines.extend(
        [
            "",
            "Inhalt dieses ZIP:",
            "- export_manifest.txt (Formatversion und Dateiliste)",
            "- datev_hinweise.txt (DATEV-Grenzen, kein DATEVconnect)",
            "- summary.txt (diese Datei)",
            "- pruefbericht_paket.txt (Prüfberichte aller exportierten Rechnungen)",
            "- Excel-Export (Übersicht + Positionen, alle Rechnungen)",
            "- DATEV-Export (eine Buchungszeile je Rechnung)",
        ]
    )
    if has_xml:
        lines.append("- original/*.xml (ursprüngliches Rechnungs-XML)")
    if has_pdf:
        lines.append("- original/*.pdf (ZUGFeRD-PDF mit eingebettetem XML)")
    if not has_xml and not has_pdf:
        lines.append("- Originaldatei fehlt in diesem Paket")
    lines.extend(
        [
            "",
            "Hinweis: Die Prüfung betrifft Schema-/Standardkonformität.",
            "Die Entscheidung über den Vorsteuerabzug liegt beim Steuerberater.",
            DATEV_LIMITATIONS,
            "",
        ]
    )
    return "\n".join(lines)


def build_batch_export_manifest(
    invoices: List[InvoiceParseResponse],
    original_names: List[str],
    has_xml: bool,
    has_pdf: bool,
) -> str:
    """Versioned inventory for a multi-invoice Steuerberater ZIP."""
    lines: List[str] = [
        "eInvoice Export-Manifest (Sammelexport)",
        "=======================================",
        "",
        f"Formatversion: {EXPORT_FORMAT_VERSION}",
        "Bei einer inkompatiblen Änderung wird die Hauptversion erhöht.",
        "1.x darf optionale ZIP-Mitglieder ergänzen; Spalten bleiben stabil.",
        "",
        f"Rechnungen: {len(invoices)}",
        "",
        "Dateien:",
        "- export_manifest.txt",
        "- datev_hinweise.txt",
        "- summary.txt",
        "- pruefbericht_paket.txt",
        "- Excel (.xlsx, Blätter Invoice / Lines / Flat)",
        "- DATEV-CSV (Buchungsstapel-kompatibel, CP1252, eine Zeile je Rechnung)",
    ]
    if has_xml:
        lines.append("- original/*.xml")
    if has_pdf:
        lines.append("- original/*.pdf")
    if original_names:
        lines.extend(["", "Originaldateien:"])
        for name in original_names:
            lines.append(f"- {name}")
    lines.extend(
        [
            "",
            "CSV/Excel-Spalten und DATEV-Felder: siehe GET /api/invoices/export/mapping",
            "und docs/EXPORT_MAPPING.md.",
            "",
            DATEV_LIMITATIONS,
            "",
        ]
    )
    return "\n".join(lines)


def build_batch_validation_report(invoices: List[InvoiceParseResponse]) -> str:
    """Concatenate per-invoice Prüfberichte for the batch ZIP."""
    parts: List[str] = []
    for invoice in invoices:
        parts.append(build_validation_report(invoice).rstrip())
    return ("\n\n-----\n\n".join(parts) + "\n") if parts else ""


def build_export_manifest(invoice: InvoiceParseResponse, has_xml: bool, has_pdf: bool) -> str:
    """Versioned inventory so Kanzlei imports stay stable across product updates."""
    lines: List[str] = [
        "eInvoice Export-Manifest",
        "========================",
        "",
        f"Formatversion: {EXPORT_FORMAT_VERSION}",
        "Bei einer inkompatiblen Änderung wird die Hauptversion erhöht.",
        "",
        f"Quelldatei: {invoice.filename}",
        f"Typ: {invoice.file_type or '—'}",
        f"Rechnung: {invoice.invoice_number or '—'}",
        f"Datum: {_de_date(invoice.issue_date) or '—'}",
        "",
        "Dateien:",
        "- export_manifest.txt",
        "- datev_hinweise.txt",
        "- summary.txt",
        "- Prüfbericht (.txt)",
        "- Excel (.xlsx, Blätter Invoice / Lines / Flat)",
        "- DATEV-CSV (Buchungsstapel-kompatibel, CP1252)",
    ]
    if has_xml:
        lines.append("- original/*.xml")
    if has_pdf:
        lines.append("- original/*.pdf")
    lines.extend(
        [
            "",
            "CSV/Excel-Spalten und DATEV-Felder: siehe GET /api/invoices/export/mapping",
            "und docs/EXPORT_MAPPING.md.",
            "",
            DATEV_LIMITATIONS,
            "",
        ]
    )
    return "\n".join(lines)


def build_datev_notes() -> str:
    """German DATEV import notes shipped with every accountant package."""
    return "\n".join(
        [
            "DATEV-Export — Hinweise",
            "=======================",
            "",
            f"Formatversion: {EXPORT_FORMAT_VERSION}",
            "",
            DATEV_LIMITATIONS,
            "",
            "Technische Form:",
            "- Trennzeichen: Semikolon",
            "- Dezimaltrennzeichen: Komma",
            "- Zeichensatz: Windows-1252 (CP1252)",
            "- Belegdatum: TTMMJJJJ",
            "- Umsatz immer positiv; Richtung über Soll (S) / Haben (H)",
            "",
            "Enthalten:",
            "- Eine Buchungszeile über den Bruttobetrag",
            "- Belegfeld 1 = Rechnungsnummer",
            "- Buchungstext = Lieferant + Rechnungsnummer",
            "",
            "Durch die Kanzlei zu ergänzen:",
            "- Beraternummer, Mandantennummer, Wirtschaftsjahr",
            "- Konto und Gegenkonto (Kontenrahmen)",
            "- BU-Schlüssel, Kostenstellen, Skonto",
            "- DATEV-EXTF-Kopfzeile, falls der Import sie verlangt",
            "",
        ]
    )


def decode_base64_payload(value: str) -> bytes:
    """Decode raw or data-URL base64 content."""
    raw: str = value.strip()
    if raw.startswith("data:") and "," in raw:
        raw = raw.split(",", 1)[1]
    try:
        return base64.b64decode(raw, validate=False)
    except (binascii.Error, ValueError) as exc:
        raise ValueError("Datei konnte nicht gelesen werden (ungültiges Base64).") from exc


def decode_pdf_base64(value: str) -> bytes:
    """Decode raw or data-URL base64 PDF content."""
    return decode_base64_payload(value)


def assert_xml_bytes(data: bytes) -> None:
    """Reject payloads that are clearly not XML."""
    stripped: bytes = data.lstrip(b"\xef\xbb\xbf \t\r\n")
    if not stripped.startswith(b"<"):
        raise ValueError("Die angehängte Datei ist kein gültiges XML.")


def _extract_xml_from_zugferd(pdf_bytes: bytes) -> Optional[bytes]:
    try:
        xml_text: Optional[str] = extract_embedded_xml_from_pdf(pdf_bytes)
    except Exception:
        return None
    if not xml_text:
        return None
    encoded: bytes = xml_text.encode("utf-8")
    try:
        assert_xml_bytes(encoded)
    except ValueError:
        return None
    return encoded


def _assert_source_size(data: Optional[bytes], label: str) -> None:
    if data is not None and len(data) > MAX_PACKAGE_SOURCE_BYTES:
        raise ValueError(
            f"{label} zu groß für das Paket (max. {MAX_PACKAGE_SOURCE_BYTES // (1024 * 1024)} MB)."
        )


def _package_source_member_name(
    invoice: InvoiceParseResponse,
    original_filename: Optional[str],
    extension: str,
) -> str:
    stem: str = ""
    if original_filename:
        base: str = original_filename.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
        dotted: str = f".{extension}"
        if base.lower().endswith(dotted):
            stem = safe_filename_stem(base[: -len(dotted)])
        else:
            stem = safe_filename_stem(base)
    if not stem:
        supplier: str = safe_filename_stem(invoice.seller.name if invoice.seller else None) or "supplier"
        number: str = safe_filename_stem(invoice.invoice_number) or "invoice"
        stem = f"{supplier}_{number}"
    return f"{ORIGINAL_DIR}/{stem}.{extension}"


def _batch_original_member_name(position: int, filename: str) -> str:
    """original/01_stem.xml — position prefix avoids collisions across the batch."""
    base: str = Path(filename).name
    suffix: str = Path(base).suffix.lower()
    if suffix not in {".xml", ".pdf"}:
        suffix = Path(base).suffix or ".dat"
    stem: str = safe_filename_stem(Path(base).stem) or "datei"
    return f"{ORIGINAL_DIR}/{position:02d}_{stem}{suffix}"


def _validation_engine_line(invoice: InvoiceParseResponse) -> str:
    meta: ValidationMeta = invoice.validation_meta
    if meta.engine == "kosit":
        version: str = f" {meta.engine_version}" if meta.engine_version else ""
        scenarios: str = f" · {meta.scenarios_version}" if meta.scenarios_version else ""
        return f"KoSIT Validator{version}{scenarios}"
    return "interne Geschäftsregeln (keine volle KoSIT-Prüfung)"


def _mismatch_summary_line(invoice: InvoiceParseResponse) -> str:
    if invoice.file_type != "zugferd_pdf" or not invoice.mismatch_fields:
        return "nicht geprüft / nicht zutreffend"
    mismatches: List[MismatchField] = [
        item for item in invoice.mismatch_fields if item.xml_value and not item.matched
    ]
    if mismatches:
        labels: str = ", ".join(item.label for item in mismatches)
        return f"Abweichung ({labels})"
    return "übereinstimmend"


def _localize_csv_row(row: Dict[str, Any]) -> Dict[str, Any]:
    localized: Dict[str, Any] = dict(row)
    for date_field in ("issue_date", "due_date"):
        localized[date_field] = _de_date(str(row.get(date_field) or "")) or row.get(date_field, "")
    for amount_field in (
        "net",
        "tax",
        "gross",
        "line_quantity",
        "line_unit_price",
        "line_tax_rate",
        "line_net_amount",
    ):
        raw: Any = row.get(amount_field, "")
        if raw == "" or raw is None:
            localized[amount_field] = ""
        else:
            localized[amount_field] = str(raw).replace(".", ",")
    return localized


def _num_str(value: Optional[Decimal]) -> str:
    if value is None:
        return ""
    return f"{value:.2f}"


def _de_amount(value: Optional[Decimal]) -> str:
    if value is None:
        return ""
    return f"{value:.2f}".replace(".", ",")


def _de_date(iso_date: Optional[str]) -> str:
    if not iso_date or len(iso_date) < 10:
        return iso_date or ""
    return f"{iso_date[8:10]}.{iso_date[5:7]}.{iso_date[0:4]}"


def _datev_date(iso_date: Optional[str]) -> str:
    """DATEV Belegdatum: DDMMYYYY."""
    if not iso_date or len(iso_date) < 10:
        return ""
    year: str = iso_date[0:4]
    month: str = iso_date[5:7]
    day: str = iso_date[8:10]
    return f"{day}{month}{year}"
